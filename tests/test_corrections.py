import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import load_workbook

from cosmo_tiling.corrections import (
    CorrectionAnalysis,
    LLMAction,
    LLMActionBatch,
    OpenAIInterpreter,
    analyze_correction,
    generate_corrected_workbook,
)
from cosmo_tiling.parsers.common import OrderRow

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "pdf" / "revised-orders"
TEMPLATE = ROOT / "src" / "cosmo_tiling" / "config" / "templates" / "classica-template.json"


class NoopInterpreter(OpenAIInterpreter):
    def interpret(self, entries, candidate_rows):
        return []


class CorrectionTests(unittest.TestCase):
    def analyze(self, original_name: str, corrected_name: str):
        with patch.dict(os.environ, {}, clear=True):
            return analyze_correction(
                FIXTURES / original_name,
                FIXTURES / corrected_name,
                TEMPLATE,
                interpreter=NoopInterpreter(),
            )

    def test_revised_header_metadata_survives_reflow(self):
        analysis, _original, corrected = self.analyze(
            "VendorOrder_33McLeanSouthShore.pdf",
            "Vendor Order_33 Mclean Southshore_Updated.pdf",
        )

        self.assertEqual(analysis.project, "33 McLean - South Shore")
        self.assertEqual(corrected.metadata["Customer"], "Brooks & Carol Johnson")
        self.assertEqual(
            corrected.metadata["Job Address"],
            "1207 Great Egret Drive, Clover, SC 29710",
        )

    def test_lot_33_detects_niche_and_corner_shelf_change(self):
        analysis, _original, _corrected = self.analyze(
            "VendorOrder_33McLeanSouthShore.pdf",
            "Vendor Order_33 Mclean Southshore_Updated.pdf",
        )

        self.assertTrue(
            any(
                action.operation == "ADD"
                and action.room == "BTHF_BR2"
                and action.item_type == "Schluter niche"
                for action in analysis.actions
            )
        )
        self.assertTrue(
            any(
                "corner shelf color" in action.evidence_corrected.casefold()
                and "CN13" in action.after_value
                for action in analysis.actions
            )
        )
        self.assertFalse(analysis.requires_review, analysis.model_dump_json(indent=2))

    def test_lot_34_preserves_unstated_quantity_impact(self):
        analysis, _original, _corrected = self.analyze(
            "VendorOrder_34McLeanSouthShore.pdf",
            "Vendor Order-34 Mclean southshore -Updated.pdf",
        )

        quantity_actions = [
            action
            for action in analysis.actions
            if action.quantity_treatment == "preserved"
        ]
        self.assertTrue(quantity_actions)
        self.assertTrue(
            any("baseline quantity is retained" in warning for action in quantity_actions for warning in action.warnings)
        )

    def test_generation_adds_revision_report(self):
        analysis, _original, _corrected = self.analyze(
            "VendorOrder_33McLeanSouthShore.pdf",
            "Vendor Order_33 Mclean Southshore_Updated.pdf",
        )
        decisions = {
            action.id: "apply"
            for action in analysis.actions
            if action.confidence == "review"
        }
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "corrected.xlsx"
            _final, resolved, rows = generate_corrected_workbook(
                FIXTURES / "VendorOrder_33McLeanSouthShore.pdf",
                FIXTURES / "Vendor Order_33 Mclean Southshore_Updated.pdf",
                output,
                TEMPLATE,
                decisions,
                expected_analysis=CorrectionAnalysis.model_validate(analysis.model_dump()),
            )
            workbook = load_workbook(output, data_only=False)
            try:
                self.assertIn("Revision Report", workbook.sheetnames)
                report = workbook["Revision Report"]
                statuses = [report.cell(row, 1).value for row in range(10, report.max_row + 1)]
                self.assertTrue(any(status in {"applied", "already_current"} for status in statuses))
            finally:
                workbook.close()
        self.assertTrue(rows)
        self.assertTrue(resolved)
        self.assertFalse(
            any(row.room == "BTHF_BR2" and row.item_type == "Corner Shelf" for row in rows)
        )
        self.assertTrue(
            any(row.room == "BTHF_BR2" and row.item_type == "Schluter niche" for row in rows)
        )
        self.assertTrue(
            any(
                row.room == "BTHF_BR3"
                and row.item_type == "Corner Shelf"
                and "CN13" in row.description
                for row in rows
            )
        )

    def test_lot_34_generation_retains_baseline_quantities_and_reports_warning(self):
        analysis, original, _corrected = self.analyze(
            "VendorOrder_34McLeanSouthShore.pdf",
            "Vendor Order-34 Mclean southshore -Updated.pdf",
        )
        decisions = {
            action.id: "apply"
            for action in analysis.actions
            if action.confidence == "review"
        }
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "corrected.xlsx"
            _final, resolved, rows = generate_corrected_workbook(
                FIXTURES / "VendorOrder_34McLeanSouthShore.pdf",
                FIXTURES / "Vendor Order-34 Mclean southshore -Updated.pdf",
                output,
                TEMPLATE,
                decisions,
                expected_analysis=analysis,
            )
            workbook = load_workbook(output, data_only=False)
            try:
                report = workbook["Revision Report"]
                warnings = [
                    str(report.cell(row, 11).value or "")
                    for row in range(10, report.max_row + 1)
                ]
                self.assertTrue(any("baseline quantity is retained" in warning for warning in warnings))
            finally:
                workbook.close()

        original_quantities = [
            (row.item_type, row.measured_qty, row.order_qty)
            for row in original.rows
            if row.room == "BTHF_PRIM"
        ]
        final_quantities = [
            (row.item_type, row.measured_qty, row.order_qty)
            for row in rows
            if row.room == "BTHF_PRIM"
        ]
        self.assertEqual(final_quantities, original_quantities)
        self.assertTrue(any(action.quantity_treatment == "preserved" for action in resolved))

    def test_rejects_swapped_documents(self):
        with self.assertRaisesRegex(ValueError, "later issue date"):
            self.analyze(
                "Vendor Order_33 Mclean Southshore_Updated.pdf",
                "VendorOrder_33McLeanSouthShore.pdf",
            )

    def test_rejects_identical_documents(self):
        name = "VendorOrder_33McLeanSouthShore.pdf"
        with self.assertRaisesRegex(ValueError, "identical"):
            self.analyze(name, name)

    def test_rejects_mismatched_projects(self):
        with self.assertRaisesRegex(ValueError, "same project"):
            self.analyze(
                "VendorOrder_33McLeanSouthShore.pdf",
                "Vendor Order-34 Mclean southshore -Updated.pdf",
            )

    @patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}, clear=True)
    @patch("openai.OpenAI")
    def test_llm_uses_structured_output_and_redacts_identifiers(self, mock_openai):
        parsed = LLMActionBatch(actions=[LLMAction(
            operation="CHANGE",
            room="BTHF_BR3",
            item_type="Corner Shelf",
            target_field="description",
            after_value="Gray CN13",
            evidence="Corner shelf color to CN13",
            candidate_index=0,
        )])
        client = Mock()
        client.responses.parse.return_value.output_parsed = parsed
        mock_openai.return_value = client

        actions = OpenAIInterpreter().interpret(
            ["Contact jane@example.com at 704-555-0100; corner shelf color to CN13"],
            [OrderRow(room="BTHF_BR3", item_type="Corner Shelf", description="White CN10")],
        )

        self.assertEqual(actions, parsed.actions)
        call = client.responses.parse.call_args
        self.assertFalse(call.kwargs["store"])
        self.assertIs(call.kwargs["text_format"], LLMActionBatch)
        prompt = call.kwargs["input"][1]["content"]
        self.assertNotIn("jane@example.com", prompt)
        self.assertNotIn("704-555-0100", prompt)

    @patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}, clear=True)
    @patch("openai.OpenAI", side_effect=RuntimeError("unavailable"))
    def test_llm_failure_returns_deterministic_fallback(self, _mock_openai):
        actions = OpenAIInterpreter().interpret(
            ["Change grout color"],
            [OrderRow(room="BTHF_PRIM", item_type="Grout")],
        )

        self.assertEqual(actions, [])

    @patch.dict(os.environ, {}, clear=True)
    @patch("openai.OpenAI")
    def test_llm_is_not_called_without_credentials(self, mock_openai):
        actions = OpenAIInterpreter().interpret(
            ["Change grout color"],
            [OrderRow(room="BTHF_PRIM", item_type="Grout")],
        )

        self.assertEqual(actions, [])
        mock_openai.assert_not_called()

    @patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}, clear=True)
    @patch("openai.OpenAI")
    def test_llm_refusal_without_parsed_output_falls_back(self, mock_openai):
        client = Mock()
        client.responses.parse.return_value.output_parsed = None
        mock_openai.return_value = client

        actions = OpenAIInterpreter().interpret(
            ["Change grout color"],
            [OrderRow(room="BTHF_PRIM", item_type="Grout")],
        )

        self.assertEqual(actions, [])

    @patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}, clear=True)
    @patch("openai.OpenAI")
    def test_llm_timeout_falls_back(self, mock_openai):
        client = Mock()
        client.responses.parse.side_effect = TimeoutError("timed out")
        mock_openai.return_value = client

        actions = OpenAIInterpreter().interpret(
            ["Change grout color"],
            [OrderRow(room="BTHF_PRIM", item_type="Grout")],
        )

        self.assertEqual(actions, [])

    @patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}, clear=True)
    @patch("openai.OpenAI")
    def test_llm_malformed_structured_output_falls_back(self, mock_openai):
        client = Mock()
        client.responses.parse.side_effect = ValueError("invalid structured output")
        mock_openai.return_value = client

        actions = OpenAIInterpreter().interpret(
            ["Change grout color"],
            [OrderRow(room="BTHF_PRIM", item_type="Grout")],
        )

        self.assertEqual(actions, [])


if __name__ == "__main__":
    unittest.main()

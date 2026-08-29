import io
import unittest
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.corrections import app

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "pdf" / "revised-orders"
ORIGINAL = FIXTURES / "VendorOrder_33McLeanSouthShore.pdf"
CORRECTED = FIXTURES / "Vendor Order_33 Mclean Southshore_Updated.pdf"


def correction_files():
    return {
        "original_pdf": (ORIGINAL.name, ORIGINAL.read_bytes(), "application/pdf"),
        "corrected_pdf": (CORRECTED.name, CORRECTED.read_bytes(), "application/pdf"),
    }


class CorrectionsApiTests(unittest.TestCase):
    @patch.dict("os.environ", {}, clear=True)
    def test_analyze_returns_deterministic_actions(self):
        response = TestClient(app).post(
            "/api/corrections/analyze",
            files=correction_files(),
            data={"template": "classica"},
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["project"], "33 McLean - South Shore")
        self.assertTrue(body["actions"])
        self.assertFalse(body["requires_review"])

    @patch("api.corrections._save_history", return_value=False)
    @patch.dict("os.environ", {}, clear=True)
    def test_generate_returns_corrected_workbook(self, _mock_history):
        client = TestClient(app)
        analyzed = client.post(
            "/api/corrections/analyze",
            files=correction_files(),
            data={"template": "classica"},
        )
        analysis = analyzed.json()
        decisions = {
            action["id"]: "apply"
            for action in analysis["actions"]
            if action["confidence"] == "review"
        }

        response = client.post(
            "/api/corrections/generate",
            files=correction_files(),
            data={
                "template": "classica",
                "analysis": __import__("json").dumps(analysis),
                "decisions": __import__("json").dumps(decisions),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("-Corrected.xlsx", response.headers["content-disposition"])
        self.assertEqual(response.headers["x-history-saved"], "false")
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        try:
            self.assertIn("Revision Report", workbook.sheetnames)
        finally:
            workbook.close()

    @patch.dict("os.environ", {}, clear=True)
    def test_rejects_identical_documents(self):
        content = ORIGINAL.read_bytes()
        response = TestClient(app).post(
            "/api/corrections/analyze",
            files={
                "original_pdf": ("one.pdf", content, "application/pdf"),
                "corrected_pdf": ("two.pdf", content, "application/pdf"),
            },
            data={"template": "classica"},
        )

        self.assertEqual(response.status_code, 422)
        self.assertIn("identical", response.json()["detail"])

    def test_saussy_corrections_are_gated(self):
        response = TestClient(app).post(
            "/api/corrections/analyze",
            files=correction_files(),
            data={"template": "saussy"},
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("not enabled", response.json()["detail"])


if __name__ == "__main__":
    unittest.main()

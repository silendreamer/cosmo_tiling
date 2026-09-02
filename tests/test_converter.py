import re
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from convert_tile_order import (
    build_reference_rows,
    build_saussy_rows,
    convert,
    extract_pdf_lines,
    extract_size,
    load_template,
    load_template_rules,
    parse_metadata,
    parse_order_rows,
    parse_saussy_metadata,
    resolve_template_rules,
)

ROOT = Path(__file__).resolve().parents[1]
SAMPLE = ROOT / "pdf" / "classica" / "VendorOrder_PalosVerdeEstates7.pdf"
PALISADES = ROOT / "pdf" / "classica" / "VendorOrder_ThePalisades3Homearama.pdf"
LAKESIDE = ROOT / "pdf" / "qa-corrections" / "VendorOrder_LakesideDriveII3.pdf"
TEMPLATE = ROOT / "src" / "cosmo_tiling" / "config" / "templates" / "classica-template.json"
SAUSSY_TEMPLATE = ROOT / "src" / "cosmo_tiling" / "config" / "templates" / "saussy-template.json"
SAUSSY_PDFS = [
    ROOT / "pdf" / "saussy" / "Eastland 104 Modern Luxe DSS 12.15.2025.pdf",
    ROOT / "pdf" / "saussy" / "ELY 125 DSS 3.16.26.pdf",
    ROOT / "pdf" / "saussy" / "Zweier DSS 11.13.2025.pdf",
]


class ConverterTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        lines, _ = extract_pdf_lines(SAMPLE)
        cls.metadata = parse_metadata(lines)
        cls.rows = parse_order_rows(lines)
        cls.order_rows, _ = build_reference_rows(
            cls.rows, load_template_rules(cls.metadata, TEMPLATE)
        )

    def test_sample_rooms_are_dynamic_and_tile_only(self):
        rooms = list(dict.fromkeys(row.room for row in self.rows))
        self.assertEqual(
            rooms,
            [
                "BTHF_BR2",
                "BTHF_BR3",
                "BTHF_BR4",
                "BTHF_OWN",
                "BTHPOWDERUP",
                "KITCHEN/inc Pantry",
                "LNDRY1",
                "LNDRY2",
            ],
        )
        self.assertFalse(any(row.source_text.upper().startswith("HARDWOOD") for row in self.rows))
        self.assertFalse(any("STRUCTURAL OPTION" in row.source_text.upper() for row in self.rows))

    def test_measured_and_waste_inclusive_quantities(self):
        floor = next(
            row
            for row in self.rows
            if row.room == "BTHF_BR2" and row.item_type == "Floor Tile"
        )
        self.assertEqual((floor.measured_qty, floor.order_qty, floor.unit), (40, 50, "SF"))
        self.assertIn("FAMED: 12x24", floor.description)

        backsplash = next(
            row
            for row in self.rows
            if row.room == "KITCHEN/inc Pantry"
            and row.item_type == "Backsplash"
        )
        self.assertEqual(
            (backsplash.measured_qty, backsplash.order_qty, backsplash.unit),
            (48, 60, "SF"),
        )

    def test_wrapped_niche_description_is_not_a_fake_corner_shelf(self):
        niche = next(
            row
            for row in self.rows
            if row.room == "BTHF_BR2" and row.item_type == "Niche"
        )
        self.assertIn("lieu of corner shelf", niche.description)
        self.assertFalse(any(row.item_type == "Corner Shelf" for row in self.rows))

    def test_niche_size_accepts_curly_inch_marks(self):
        description = (
            "Wall Niche / Schluter / KB-12-SN-305-711-A1 / Double niche "
            "w/shelf 28” x 12” (2 – 12” cavities)"
        )
        self.assertEqual(extract_size(description), '28"x12"')

    def test_generated_workbook_has_grouped_and_flat_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "sample.xlsx"
            rows = convert(SAMPLE, output, template_path=TEMPLATE)
            workbook = load_workbook(output, read_only=True, data_only=False)
            try:
                self.assertEqual(workbook.sheetnames, ["Tile Order", "Data"])
                self.assertEqual(workbook["Data"].max_row - 1, len(rows))
                self.assertEqual(workbook["Tile Order"].max_column, 7)
                self.assertEqual(
                    [workbook["Tile Order"].cell(14, column).value for column in range(1, 8)],
                    [
                        "Type", "Size / Area", "Description", "Order Qty",
                        "Unit", "Comments", "Pattern",
                    ],
                )
                self.assertEqual(workbook["Data"]["D1"].value, "Measured Qty")

                owner_wall = next(
                    row
                    for row in workbook["Tile Order"].iter_rows(min_row=15)
                    if row[0].value == "Shower wall"
                    and row[2].value == "CONRAD BRICK, Siren, CB95"
                )
                self.assertIsInstance(owner_wall[3].value, str)
                self.assertTrue(owner_wall[3].value.startswith("=ROUND((235)"))
                self.assertIsNone(re.search(r"\b[A-Z]{1,3}\d+\b", owner_wall[3].value))
            finally:
                workbook.close()

    def test_reference_style_rows_and_rules(self):
        self.assertEqual(len(self.order_rows), 56)
        self.assertEqual(
            list(dict.fromkeys(row.room for row in self.order_rows)),
            [
                "BTHF_OWN", "BTHF_BR2", "BTHF_BR3", "BTHF_BR4",
                "KITCHEN/inc Pantry", "LNDRY1", "LNDRY2", "BTHPOWDERUP",
            ],
        )
        owner_wall = next(
            row
            for row in self.order_rows
            if row.room == "BTHF_OWN" and row.item_type == "Shower wall"
        )
        self.assertEqual(owner_wall.description, "CONRAD BRICK, Siren, CB95")
        self.assertEqual(owner_wall.measured_qty, 235)
        self.assertEqual(owner_wall.waste_percent, 10)
        self.assertEqual(
            owner_wall.pattern,
            'Stacked Vertical (less than 15" in length)',
        )

        combined_floor = next(
            row
            for row in self.order_rows
            if row.room == "BTHF_BR2" and row.item_type == "Floor Tile"
        )
        self.assertEqual(combined_floor.order_components, (39, 37))
        self.assertEqual(combined_floor.waste_percent, 18)
        self.assertIn("BTHF_BR4", combined_floor.comments)

        self.assertTrue(
            any(row.item_type == "Drain riser plug" for row in self.order_rows)
        )
        self.assertTrue(any(row.item_type == "Caulk" for row in self.order_rows))

    def test_palisades_standard_and_double_niche_sizes(self):
        lines, _ = extract_pdf_lines(PALISADES)
        metadata = parse_metadata(lines)
        rows, _ = build_reference_rows(
            parse_order_rows(lines), load_template_rules(metadata, TEMPLATE)
        )
        standard = next(
            row
            for row in rows
            if row.room == "BTHF_BR2" and row.item_type == "Schluter niche"
        )
        double = next(
            row
            for row in rows
            if row.room == "BTHF_OWN" and row.item_type == "Schluter niche"
        )
        self.assertEqual(standard.size_area, '12"x12"')
        self.assertEqual(double.size_area, '28"x12"')

    def test_lakeside_qa_corrections_are_applied(self):
        lines, _ = extract_pdf_lines(LAKESIDE)
        metadata = parse_metadata(lines)
        raw_rows = parse_order_rows(lines)

        self.assertEqual(
            list(dict.fromkeys(row.room for row in raw_rows)),
            [
                "BTHF_BR2", "BTHF_BR3", "BTHF_BR4S", "BTHF_PRIM",
                "GREAT", "KITCHEN", "LNDRY1", "SCULLERY",
            ],
        )
        raw_primary_wall = next(
            row for row in raw_rows
            if row.room == "BTHF_PRIM" and row.item_type == "Shower Wall"
        )
        self.assertIn("COSTAR", raw_primary_wall.description)
        self.assertNotIn("MYTHIQUE", raw_primary_wall.source_text)

        rows, display_names = build_reference_rows(
            raw_rows, load_template_rules(metadata, TEMPLATE)
        )
        self.assertEqual(
            list(dict.fromkeys(row.room for row in rows)),
            [
                "BTHF_BR2", "BTHF_BR3", "BTHF_BR4S", "BTHF_PRIM",
                "GREAT", "LNDRY1", "SCULLERY",
            ],
        )
        self.assertEqual(display_names["GREAT"], "Great Room Fireplace")
        self.assertEqual(display_names["SCULLERY"], "Scullery Backsplash")

        tub_caulk = [
            row for row in rows
            if row.room == "BTHF_BR2" and row.item_type == "Caulk"
        ]
        self.assertEqual(len(tub_caulk), 2)
        self.assertEqual(tub_caulk[0].description, "Ash (642) Sanded")
        self.assertEqual((tub_caulk[0].order_qty, tub_caulk[0].unit), (1, "PCS"))
        self.assertEqual((tub_caulk[1].order_qty, tub_caulk[1].unit), ("-", "-"))

        sealer_rooms = [row.room for row in rows if row.item_type == "Sealer"]
        self.assertEqual(sealer_rooms, ["BTHF_BR3", "BTHF_BR4S", "GREAT", "SCULLERY"])
        self.assertTrue(
            all((row.order_qty, row.unit) == (1, "QT") for row in rows if row.item_type == "Sealer")
        )

        grouped_wall = next(
            row for row in rows
            if row.room == "BTHF_BR2" and row.item_type == "Shower wall"
        )
        self.assertEqual(grouped_wall.order_components, (81, 108, 108))
        self.assertEqual(grouped_wall.waste_percent, 22)
        self.assertEqual(grouped_wall.comments, "=81+108+108")

        primary_wall = next(
            row for row in rows
            if row.room == "BTHF_PRIM" and row.item_type == "Shower wall"
        )
        self.assertEqual(
            (primary_wall.size_area, primary_wall.description),
            ("12x24", "COSTAR, Oyster, CT74"),
        )
        self.assertEqual((primary_wall.measured_qty, primary_wall.waste_percent), (153, 22))
        self.assertTrue(
            any(row.room == "BTHF_PRIM" and row.item_type == "Drain riser plug" for row in rows)
        )

        fireplace = next(
            row for row in rows if row.room == "GREAT" and row.item_type == "Tile"
        )
        self.assertEqual(fireplace.description, "MYTHIQUE MARBLE, Majestic, Polished, MY12")
        self.assertEqual((fireplace.measured_qty, fireplace.waste_percent), (143, 22))
        fireplace_schluter = next(
            row for row in rows if row.room == "GREAT" and row.item_type == "Schluter"
        )
        self.assertEqual((fireplace_schluter.description, fireplace_schluter.order_qty), ("J100-PG (classic grey)", 4))

        scullery = next(
            row for row in rows if row.room == "SCULLERY" and row.item_type == "Tile"
        )
        self.assertEqual(scullery.description, "SUBLIMITY, Daphne White, Balance Mosaic, M103")
        self.assertEqual((scullery.measured_qty, scullery.waste_percent), (13, 20))
        self.assertEqual(
            scullery.pattern,
            'Stacked Horizontal (up to 24" length): Parallel to Cabinet',
        )

    def test_classica_template_selects_matching_project_rules(self):
        rules = load_template_rules(self.metadata, TEMPLATE)
        self.assertEqual(rules["measurements"]["BTHF_OWN|Shower Wall"], 235)
        self.assertEqual(rules["waste_percent"]["BTHF_OWN|Shower Wall"], 10)

        palisades_lines, _ = extract_pdf_lines(PALISADES)
        palisades_rules = load_template_rules(
            parse_metadata(palisades_lines), TEMPLATE
        )
        self.assertEqual(palisades_rules, {})

    def test_saussy_template_selects_all_reference_profiles(self):
        template = load_template(SAUSSY_TEMPLATE)
        expected = [
            ("Central Villages at Eastland Yards Lot 104", 22, 3),
            ("Eastland Yards Lot 125", 22, 3),
            ("Zweier", 35, 6),
        ]
        for pdf, (project_fragment, row_count, room_count) in zip(SAUSSY_PDFS, expected):
            lines, _ = extract_pdf_lines(pdf)
            metadata = parse_saussy_metadata(lines)
            self.assertIn(project_fragment.casefold(), metadata["Project"].casefold())
            rules = resolve_template_rules(template, metadata)
            rows, _ = build_saussy_rows(lines, rules)
            self.assertEqual(len(rows), row_count)
            self.assertEqual(len(dict.fromkeys(row.room for row in rows)), room_count)

    def test_unmatched_saussy_project_is_parsed_from_its_tile_section(self):
        template = load_template(SAUSSY_TEMPLATE)
        lines, _ = extract_pdf_lines(SAUSSY_PDFS[0])
        rules = resolve_template_rules(template, {"Project": "A brand-new project"})

        self.assertNotIn("rows", rules)
        rows, _ = build_saussy_rows(lines, rules)

        self.assertGreaterEqual(len(rows), 7)
        self.assertTrue(all(row.measured_qty is None for row in rows))
        self.assertTrue(
            any(
                row.room == "Owner's Bath"
                and row.item_type == "Shower wall"
                and row.description == "Serendra Mila"
                for row in rows
            )
        )
        self.assertTrue(
            any(
                row.room == "Kitchen"
                and row.item_type == "KBS Tile"
                and row.description == "Catch Gloss White"
                for row in rows
            )
        )

    def test_saussy_workbook_matches_updated_format_and_keeps_audit_data(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "saussy.xlsx"
            rows = convert(SAUSSY_PDFS[0], output, template_path=SAUSSY_TEMPLATE)
            workbook = load_workbook(output, read_only=True, data_only=False)
            try:
                order = workbook["Tile Order"]
                data = workbook["Data"]
                self.assertEqual(order.max_column, 7)
                self.assertEqual(order["A1"].value, "SAUSSY TILE ORDER")
                self.assertEqual(data["D1"].value, "Measured Qty")
                self.assertEqual(data.max_row - 1, len(rows))

                owner_wall = next(
                    row for row in order.iter_rows()
                    if row[0].value == "Shower wall"
                    and row[2].value == "Serendra Mila"
                )
                self.assertEqual(owner_wall[1].value, "12x24")
                self.assertEqual(owner_wall[3].value, "=ROUND((187)*(1+12/100),0)")
                self.assertEqual(owner_wall[6].value, "Straight Vertical")

                grout = next(
                    row for row in order.iter_rows()
                    if row[0].value == "Grout"
                    and row[2].value == "Prism #165 Delorean Gray"
                    and row[3].value is not None
                )
                self.assertEqual(grout[3].value, "=ROUND((187+46+58)/150,0)")

                audit_wall = next(
                    row for row in data.iter_rows(min_row=2)
                    if row[0].value == "Shower wall"
                    and row[2].value == "Serendra Mila"
                )
                self.assertEqual(audit_wall[3].value, 187)
                self.assertIn("Saussy Design Selection Sheet", audit_wall[9].value)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()

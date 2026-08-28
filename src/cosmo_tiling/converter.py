from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from copy import deepcopy
from pathlib import Path
from typing import Iterable

from cosmo_tiling.parsers.common import OrderRow, clean_text
from cosmo_tiling.parsers.classica import (
    FOOTER_RE,
    build_reference_rows,
    extract_size,
    parse_metadata,
    parse_order_rows,
)
from cosmo_tiling.parsers.saussy import build_saussy_rows, parse_saussy_metadata

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PACKAGE_ROOT = Path(__file__).resolve().parent
DEFAULT_TEMPLATE = PACKAGE_ROOT / "config" / "templates" / "classica-template.json"


def extract_pdf_lines(pdf_path: Path) -> tuple[list[str], str]:
    lines: list[str] = []
    pages: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            pages.append(text)
            for raw_line in text.splitlines():
                line = clean_text(raw_line)
                if line and not FOOTER_RE.match(line):
                    lines.append(line)
    return lines, "\n\n--- PAGE BREAK ---\n\n".join(pages)


def merge_rules(base: dict, override: dict) -> dict:
    result = deepcopy(base)
    for key, value in override.items():
        if key == "project_match":
            continue
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = merge_rules(result[key], value)
        else:
            result[key] = deepcopy(value)
    return result


def load_template(template_path: Path | str = DEFAULT_TEMPLATE) -> dict:
    path = Path(template_path).resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Template not found: {path}")
    try:
        template = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        raise ValueError(f"Invalid template JSON at {path}: {error}") from error

    if template.get("schema_version") != 1:
        raise ValueError(
            f"Unsupported template schema_version in {path}: "
            f"{template.get('schema_version')!r}"
        )
    if template.get("parser") not in {"classica", "saussy"}:
        raise ValueError(
            f"Unsupported template parser in {path}: {template.get('parser')!r}"
        )
    if not isinstance(template.get("defaults", {}), dict):
        raise ValueError(f"Template defaults must be an object: {path}")
    if not isinstance(template.get("projects", []), list):
        raise ValueError(f"Template projects must be an array: {path}")

    reference_rules = template.get("reference_rules")
    if reference_rules:
        reference_path = (path.parent / reference_rules).resolve()
        if not reference_path.is_file():
            raise FileNotFoundError(f"Reference rules not found: {reference_path}")
        try:
            reference = json.loads(reference_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as error:
            raise ValueError(
                f"Invalid reference-rules JSON at {reference_path}: {error}"
            ) from error
        projects = reference.get("projects", [])
        if not isinstance(projects, list):
            raise ValueError(f"Reference-rule projects must be an array: {reference_path}")
        template["projects"] = projects

    return template


def resolve_template_rules(
    template: dict, metadata: OrderedDict[str, str]
) -> dict:
    rules = deepcopy(template.get("defaults", {}))
    project = metadata.get("Project", "").casefold()
    for project_rules in template.get("projects", []):
        project_match = project_rules.get("project_match", "").casefold()
        if project_match and project_match in project:
            rules = merge_rules(rules, project_rules)
            break
    return rules


def load_template_rules(
    metadata: OrderedDict[str, str], template_path: Path | str = DEFAULT_TEMPLATE
) -> dict:
    return resolve_template_rules(load_template(template_path), metadata)



DISPLAY_HEADERS = [
    "Type", "Size / Area", "Description", "Order Qty", "Unit", "Comments",
    "Pattern",
]
DATA_HEADERS = [
    "Type", "Size / Area", "Description", "Measured Qty", "Order Qty",
    "Unit", "Comments", "Pattern", "Room Code", "Source Text",
]


def order_formula(row: OrderRow, measured_cell: str | None = None) -> object:
    if row.order_formula_override:
        return row.order_formula_override
    if row.waste_percent is None or not row.order_components:
        return row.order_qty
    if len(row.order_components) == 1 and measured_cell:
        base = measured_cell
    else:
        base = f"({'+'.join(str(value) for value in row.order_components)})"
    return f"=ROUND({base}*(1+{row.waste_percent}/100),0)"


def display_row_values(row: OrderRow) -> list[object]:
    return [
        row.item_type, row.size_area, row.description, order_formula(row),
        row.unit, row.comments, row.pattern,
    ]


def data_row_values(row: OrderRow, excel_row: int) -> list[object]:
    return [
        row.item_type, row.size_area, row.description, row.measured_qty,
        order_formula(row, f"D{excel_row}"), row.unit, row.comments, row.pattern,
        row.room, row.source_text,
    ]


def style_header(cells: Iterable) -> None:
    fill = PatternFill("solid", fgColor="E7E6E6")
    thin = Side(style="thin", color="A6A6A6")
    for cell in cells:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(
    metadata: OrderedDict[str, str],
    rows: list[OrderRow],
    output_path: Path,
    display_names: dict[str, str] | None = None,
    workbook_title: str = "TILE ORDER",
    header_title: str = "Tile Order",
) -> None:
    display_names = display_names or {}
    workbook = Workbook()
    order_sheet = workbook.active
    order_sheet.title = "Tile Order"

    order_sheet.merge_cells("A1:G1")
    title = order_sheet["A1"]
    title.value = workbook_title
    title.font = Font(bold=True, size=16, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="204070")
    title.alignment = Alignment(horizontal="center")

    excel_row = 3
    for key, value in metadata.items():
        order_sheet.cell(excel_row, 1, key).font = Font(bold=True)
        order_sheet.cell(excel_row, 2, value)
        order_sheet.merge_cells(start_row=excel_row, start_column=2, end_row=excel_row, end_column=7)
        excel_row += 1
    excel_row += 1

    for column, header in enumerate(DISPLAY_HEADERS, 1):
        order_sheet.cell(excel_row, column, header)
    style_header(order_sheet[excel_row])
    excel_row += 2

    rooms: OrderedDict[str, list[OrderRow]] = OrderedDict()
    for row in rows:
        rooms.setdefault(row.room, []).append(row)

    thin = Side(style="thin", color="A6A6A6")
    for room, room_rows in rooms.items():
        order_sheet.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=7)
        room_cell = order_sheet.cell(excel_row, 1, room)
        room_cell.value = display_names.get(room, room)
        room_cell.font = Font(bold=True, size=12)
        room_cell.fill = PatternFill("solid", fgColor="D9EAD3")
        excel_row += 1

        for item in room_rows:
            for column, value in enumerate(display_row_values(item), 1):
                cell = order_sheet.cell(excel_row, column, value)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            excel_row += 1
        excel_row += 1

    widths = [22, 13, 48, 13, 9, 30, 45]
    for column, width in enumerate(widths, 1):
        order_sheet.column_dimensions[get_column_letter(column)].width = width
    order_sheet.freeze_panes = f"A{len(metadata) + 5}"
    order_sheet.sheet_view.showGridLines = False
    order_sheet.page_setup.orientation = "landscape"
    order_sheet.page_setup.fitToWidth = 1
    order_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    order_sheet.oddHeader.left.text = "DATE:  &D"
    order_sheet.oddHeader.center.text = f"{header_title} - {metadata.get('Project', '')}"

    data_sheet = workbook.create_sheet("Data")
    data_sheet.append(DATA_HEADERS)
    style_header(data_sheet[1])
    for data_row, item in enumerate(rows, 2):
        data_sheet.append(data_row_values(item, data_row))
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = data_sheet.dimensions
    data_widths = [22, 13, 48, 14, 13, 9, 30, 45, 20, 65]
    for column, width in enumerate(data_widths, 1):
        data_sheet.column_dimensions[get_column_letter(column)].width = width
    for row in data_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def validate_workbook(output_path: Path, expected_rows: int) -> None:
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        if {"Tile Order", "Data"} - set(workbook.sheetnames):
            raise RuntimeError("Generated workbook is missing a required sheet")
        data = workbook["Data"]
        actual_rows = data.max_row - 1
        if actual_rows != expected_rows:
            raise RuntimeError(f"Expected {expected_rows} data rows, found {actual_rows}")
    finally:
        workbook.close()


def convert(
    pdf_path: Path,
    output_path: Path,
    debug_text: bool = False,
    template_path: Path | str = DEFAULT_TEMPLATE,
) -> list[OrderRow]:
    lines, extracted_text = extract_pdf_lines(pdf_path)
    template = load_template(template_path)
    if template["parser"] == "classica":
        metadata = parse_metadata(lines)
        raw_rows = parse_order_rows(lines)
        if not raw_rows:
            raise ValueError("No tile-related order rows were found in the PDF")
        rules = resolve_template_rules(template, metadata)
        rows, display_names = build_reference_rows(raw_rows, rules)
    elif template["parser"] == "saussy":
        metadata = parse_saussy_metadata(lines)
        rules = resolve_template_rules(template, metadata)
        rows, display_names = build_saussy_rows(lines, rules)
    else:  # Guarded by load_template; retained for type-safe exhaustiveness.
        raise ValueError(f"Unsupported parser: {template['parser']}")
    write_workbook(
        metadata,
        rows,
        output_path,
        display_names,
        template.get("workbook_title", "TILE ORDER"),
        template.get("header_title", "Tile Order"),
    )
    validate_workbook(output_path, len(rows))
    if debug_text:
        output_path.with_suffix(".extracted.txt").write_text(extracted_text, encoding="utf-8")
    return rows


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert a vendor-order PDF into a structured tile-order workbook using a JSON template."
    )
    parser.add_argument("pdf", type=Path, help="Input vendor-order PDF")
    parser.add_argument("-o", "--output", type=Path, help="Output .xlsx path")
    parser.add_argument(
        "-t",
        "--template",
        type=Path,
        default=DEFAULT_TEMPLATE,
        help="JSON template describing the PDF format and business rules",
    )
    parser.add_argument("--debug-text", action="store_true", help="Keep pdfplumber's extracted text")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    pdf_path = args.pdf.resolve()
    if not pdf_path.is_file():
        raise SystemExit(f"PDF not found: {pdf_path}")
    output_path = (
        args.output.resolve()
        if args.output
        else Path.cwd() / "output" / f"{pdf_path.stem}-TileOrder.xlsx"
    )
    rows = convert(pdf_path, output_path, args.debug_text, args.template)

    measured_and_order = sum(
        row.measured_qty is not None
        and (row.order_qty is not None or row.waste_percent is not None)
        for row in rows
    )
    rooms = list(dict.fromkeys(row.room for row in rows))
    print(f"Created: {output_path}")
    print(f"Rooms: {len(rooms)} | Rows: {len(rows)} | Measured/order pairs: {measured_and_order}")


if __name__ == "__main__":
    main()

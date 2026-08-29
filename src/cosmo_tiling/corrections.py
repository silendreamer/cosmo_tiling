from __future__ import annotations

import hashlib
import json
import os
import re
from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from cosmo_tiling.converter import (
    extract_pdf_lines,
    load_template,
    resolve_template_rules,
    write_workbook,
)
from cosmo_tiling.parsers.classica import (
    ROOM_RE,
    build_reference_rows,
    normalize_room,
    parse_metadata,
    parse_order_rows,
    split_quantity,
)
from cosmo_tiling.parsers.common import OrderRow

PROMPT_VERSION = "correction-v1"
DEFAULT_MODEL = "gpt-5.6-terra"
ACTION_FIELDS = (
    "size_area",
    "description",
    "measured_qty",
    "order_qty",
    "unit",
    "comments",
    "pattern",
)
QUANTITY_FIELDS = {"measured_qty", "order_qty"}


class RevisionAction(BaseModel):
    id: str
    operation: Literal["ADD", "DELETE", "CHANGE", "CLARIFICATION"]
    room: str = ""
    item_type: str = ""
    target_field: str = "row"
    before_value: str = ""
    after_value: str = ""
    evidence_original: str = ""
    evidence_corrected: str = ""
    confidence: Literal["high", "review"] = "review"
    quantity_treatment: Literal[
        "not_applicable", "explicit", "preserved", "unchanged"
    ] = "not_applicable"
    status: Literal[
        "automatic", "review", "applied", "ignored", "unresolved", "already_current"
    ] = "review"
    warnings: list[str] = Field(default_factory=list)
    source: Literal["deterministic", "llm"] = "deterministic"
    target_group: str = ""
    target_occurrence: int = 0
    field_changes: dict[str, Any] = Field(default_factory=dict)
    row_payload: dict[str, Any] = Field(default_factory=dict)


class CorrectionAnalysis(BaseModel):
    original_hash: str
    corrected_hash: str
    template: str
    project: str
    permit_number: str
    original_date: str
    corrected_date: str
    actions: list[RevisionAction]
    warnings: list[str] = Field(default_factory=list)
    requires_review: bool = False
    model: str = ""
    prompt_version: str = PROMPT_VERSION


class LLMAction(BaseModel):
    operation: Literal["ADD", "DELETE", "CHANGE", "CLARIFICATION"]
    room: str = ""
    item_type: str = ""
    target_field: Literal[
        "row", "size_area", "description", "comments", "pattern", "unit",
        "measured_qty", "order_qty"
    ] = "row"
    before_value: str = ""
    after_value: str = ""
    evidence: str
    candidate_index: int | None = None
    affects_quantity: bool = False


class LLMActionBatch(BaseModel):
    actions: list[LLMAction] = Field(default_factory=list)


@dataclass
class ParsedCorrectionDocument:
    path: Path
    lines: list[str]
    metadata: dict[str, str]
    rows: list[OrderRow]
    display_names: dict[str, str]
    change_entries: list[str]


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _parse_date(value: str) -> date | None:
    match = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})", value.strip())
    if not match:
        return None
    month, day, year = (int(part) for part in match.groups())
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        pass
    return None


def _permit_identity(value: str) -> str:
    return value.split("Issued:", 1)[0].strip().casefold()


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def _row_payload(row: OrderRow) -> dict[str, Any]:
    return {
        "room": row.room,
        "item_type": row.item_type,
        "size_area": row.size_area,
        "description": row.description,
        "measured_qty": row.measured_qty,
        "order_qty": row.order_qty,
        "unit": row.unit,
        "comments": row.comments,
        "pattern": row.pattern,
        "source_text": row.source_text,
        "waste_percent": row.waste_percent,
        "order_components": list(row.order_components),
        "order_formula_override": row.order_formula_override,
        "related_type": row.related_type,
    }


def _row_from_payload(payload: dict[str, Any]) -> OrderRow:
    values = dict(payload)
    values["order_components"] = tuple(values.get("order_components") or ())
    return OrderRow(**values)


def _group_key(row: OrderRow) -> str:
    return f"{row.room}|{row.item_type}"


def _action_id(*parts: object) -> str:
    material = "\x1f".join(_display_value(part) for part in parts)
    return hashlib.sha256(material.encode("utf-8")).hexdigest()[:16]


def _extract_change_entries(lines: list[str]) -> list[str]:
    start = next(
        (
            index
            for index, line in enumerate(lines)
            if line == "Change Orders Approved" or line.startswith("Change Order #")
        ),
        -1,
    )
    if start < 0:
        return []
    entries: list[str] = []
    current = ""
    for line in lines[start:]:
        if line.startswith("Included at Start"):
            break
        if line == "Change Orders Approved" or line.startswith("Change Order #"):
            continue
        is_entry = bool(
            re.match(r'^["•]\s*', line)
            or re.search(r"\s-\s(?:ADD|DELETE|CHANGE|CLARIFICATION)\s-\s", line, re.IGNORECASE)
        )
        if is_entry:
            if current:
                entries.append(current)
            current = re.sub(r'^["•]\s*', "", line).strip()
        elif current:
            current = f"{current} {line}".strip()
    if current:
        entries.append(current)
    return entries


def _extract_corner_shelves(lines: list[str]) -> list[OrderRow]:
    shelves: list[OrderRow] = []
    room = ""
    current: OrderRow | None = None
    started = False
    for raw_line in lines:
        if raw_line == "Selection":
            started = True
            continue
        if not started:
            continue
        if raw_line == "Change Orders Approved" or raw_line.startswith("Included at Start"):
            break
        body, quantity, unit = split_quantity(raw_line)
        room_match = ROOM_RE.match(body)
        if room_match:
            next_room = normalize_room(room_match.group("room"))
            if current and current.room != next_room:
                current = None
            room = next_room
            body = (room_match.group("description") or "").strip()
        if not room:
            continue
        if body.casefold().startswith("corner shelf / large"):
            current = OrderRow(
                room=room,
                item_type="Corner Shelf",
                description=body,
                order_qty=quantity,
                unit=unit,
                source_text=raw_line,
            )
            shelves.append(current)
            continue
        if current and body.casefold().startswith("contemporary corner shelf"):
            current.description = body
            current.source_text = f"{current.source_text} {raw_line}"
            continue
        if current and body.casefold().startswith("contemporary shelf:"):
            detail = re.sub(
                r"^CONTEMPORARY SHELF:\s*(?:Group\s+\d+:\s*)?",
                "",
                body,
                flags=re.IGNORECASE,
            )
            current.description = f"{current.description} - {detail}".strip(" -")
            current.source_text = f"{current.source_text} {raw_line}"
            current = None
    return shelves


def _parse_document(path: Path, template_path: Path) -> ParsedCorrectionDocument:
    lines, _text = extract_pdf_lines(path)
    template = load_template(template_path)
    if template["parser"] != "classica":
        raise ValueError(
            "Corrected Saussy orders are not enabled until a real revised-order fixture is available."
        )
    metadata = dict(parse_metadata(lines))
    raw_rows = parse_order_rows(lines)
    if not raw_rows:
        raise ValueError(f"No tile-related order rows were found in {path.name}.")
    rules = resolve_template_rules(template, metadata)
    rows, display_names = build_reference_rows(raw_rows, rules)
    existing_shelves = {(row.room, row.description) for row in rows if row.item_type == "Corner Shelf"}
    rows.extend(
        shelf
        for shelf in _extract_corner_shelves(lines)
        if (shelf.room, shelf.description) not in existing_shelves
    )
    return ParsedCorrectionDocument(
        path=path,
        lines=lines,
        metadata=metadata,
        rows=rows,
        display_names=display_names,
        change_entries=_extract_change_entries(lines),
    )


def _validate_pair(
    original: ParsedCorrectionDocument,
    corrected: ParsedCorrectionDocument,
) -> None:
    if _sha256(original.path) == _sha256(corrected.path):
        raise ValueError("The original and corrected PDFs are identical.")

    original_project = original.metadata.get("Project", "").strip().casefold()
    corrected_project = corrected.metadata.get("Project", "").strip().casefold()
    if not original_project or original_project != corrected_project:
        raise ValueError("The PDFs do not belong to the same project/order.")

    original_permit = _permit_identity(original.metadata.get("Permit Number", ""))
    corrected_permit = _permit_identity(corrected.metadata.get("Permit Number", ""))
    if original_permit and corrected_permit and original_permit != corrected_permit:
        raise ValueError("The PDFs have different permit numbers.")

    original_date = _parse_date(original.metadata.get("Order Date", ""))
    corrected_date = _parse_date(corrected.metadata.get("Order Date", ""))
    if original_date and corrected_date and corrected_date <= original_date:
        raise ValueError("The corrected PDF must have a later issue date than the original PDF.")

    if corrected.metadata.get("Document", "").casefold() != "revised vendor order":
        raise ValueError("The corrected PDF is not marked as a revised vendor order.")


def _entry_tokens(entry: str) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-z0-9]+", entry.casefold())
        if len(token) > 2 and token not in {"tile", "order", "same", "with"}
    }


def _best_entry(entries: list[str], row: OrderRow) -> str:
    row_tokens = _entry_tokens(f"{row.room} {row.item_type} {row.description}")
    compatible = []
    for entry in entries:
        entry_room = _room_from_entry(entry)
        entry_item, _field = _item_from_entry(entry)
        if entry_room and entry_room != row.room:
            continue
        if (
            entry_item
            and entry_item.casefold() != row.item_type.casefold()
            and not (entry_item == "Schluter niche" and row.item_type == "Schluter niche")
        ):
            continue
        compatible.append(entry)
    ranked = sorted(
        compatible,
        key=lambda entry: len(row_tokens & _entry_tokens(entry)),
        reverse=True,
    )
    if not ranked:
        return ""
    overlap = len(row_tokens & _entry_tokens(ranked[0]))
    return ranked[0] if overlap >= 2 else ""


def _compare_rows(
    original_rows: list[OrderRow],
    corrected_rows: list[OrderRow],
    change_entries: list[str],
) -> tuple[list[RevisionAction], set[str]]:
    original_groups: dict[str, list[OrderRow]] = defaultdict(list)
    corrected_groups: dict[str, list[OrderRow]] = defaultdict(list)
    for row in original_rows:
        original_groups[_group_key(row)].append(row)
    for row in corrected_rows:
        corrected_groups[_group_key(row)].append(row)

    actions: list[RevisionAction] = []
    used_entries: set[str] = set()
    for group in sorted(set(original_groups) | set(corrected_groups)):
        before_rows = original_groups[group]
        after_rows = corrected_groups[group]
        room, _item_type = group.split("|", 1)
        for occurrence in range(max(len(before_rows), len(after_rows))):
            before = before_rows[occurrence] if occurrence < len(before_rows) else None
            after = after_rows[occurrence] if occurrence < len(after_rows) else None
            evidence = _best_entry(change_entries, after or before)  # type: ignore[arg-type]
            evidence_operation = _operation_from_entry(evidence) if evidence else ""
            confidence: Literal["high", "review"] = "high" if evidence else "review"

            if before is None and after is not None:
                if evidence and evidence_operation != "ADD":
                    # A parser-visible row appearing in the revised document is not
                    # proof of an addition when the author's instruction says
                    # clarification/change. Leave that prose unmatched for the
                    # semantic pass and keep this structural difference reviewable.
                    evidence = ""
                    confidence = "review"
                if evidence:
                    used_entries.add(evidence)
                actions.append(RevisionAction(
                    id=_action_id("ADD", group, occurrence, after.source_text),
                    operation="ADD",
                    room=room,
                    item_type=after.item_type,
                    after_value=after.description or after.item_type,
                    evidence_corrected=evidence or after.source_text,
                    confidence=confidence,
                    quantity_treatment=(
                        "explicit" if after.measured_qty is not None or after.order_qty is not None
                        else "not_applicable"
                    ),
                    status="automatic" if confidence == "high" else "review",
                    target_group=group,
                    target_occurrence=occurrence,
                    row_payload=_row_payload(after),
                ))
                continue
            if after is None and before is not None:
                if evidence and evidence_operation != "DELETE":
                    # Revised PDFs frequently omit or reflow a row while clarifying
                    # its material. Only explicit DELETE wording may corroborate a
                    # deletion; otherwise preserve the baseline row and interpret
                    # the correction text as a clarification below.
                    continue
                if evidence:
                    used_entries.add(evidence)
                actions.append(RevisionAction(
                    id=_action_id("DELETE", group, occurrence, before.source_text),
                    operation="DELETE",
                    room=room,
                    item_type=before.item_type,
                    before_value=before.description or before.item_type,
                    evidence_original=before.source_text,
                    evidence_corrected=evidence,
                    confidence=confidence,
                    quantity_treatment="not_applicable",
                    status="automatic" if confidence == "high" else "review",
                    target_group=group,
                    target_occurrence=occurrence,
                ))
                continue
            if before is None or after is None:
                continue

            changes = {
                field: getattr(after, field)
                for field in ACTION_FIELDS
                if getattr(before, field) != getattr(after, field)
            }
            if not changes:
                continue
            if evidence:
                used_entries.add(evidence)
            changed_fields = list(changes)
            quantity_changed = bool(QUANTITY_FIELDS & set(changed_fields))
            actions.append(RevisionAction(
                id=_action_id("CHANGE", group, occurrence, changed_fields, changes),
                operation="CHANGE",
                room=room,
                item_type=after.item_type,
                target_field=changed_fields[0] if len(changed_fields) == 1 else "row",
                before_value="; ".join(
                    f"{field}={_display_value(getattr(before, field))}" for field in changed_fields
                ),
                after_value="; ".join(
                    f"{field}={_display_value(getattr(after, field))}" for field in changed_fields
                ),
                evidence_original=before.source_text,
                evidence_corrected=evidence or after.source_text,
                confidence=confidence,
                quantity_treatment="explicit" if quantity_changed else "unchanged",
                status="automatic" if confidence == "high" else "review",
                target_group=group,
                target_occurrence=occurrence,
                field_changes=changes,
            ))
    return actions, used_entries


def _room_from_entry(entry: str) -> str:
    value = entry.casefold()
    mappings = (
        (r"bedroom\s*#?\s*2.*bath", "BTHF_BR2"),
        (r"bedroom\s*#?\s*3.*bath", "BTHF_BR3"),
        (r"bedroom\s*#?\s*4.*bath", "BTHF_BR4"),
        (r"primary bath", "BTHF_PRIM"),
        (r"powder", "BTHPOWDER"),
        (r"kitchen", "KITCHEN"),
    )
    for pattern, room in mappings:
        if re.search(pattern, value):
            return room
    return ""


def _item_from_entry(entry: str) -> tuple[str, str]:
    value = entry.casefold()
    if "wall niche" in value or "shampoo niche" in value or "inside of niche" in value:
        return "Schluter niche", "description"
    if "corner shelf" in value:
        return "Corner Shelf", "description"
    if "grout color" in value:
        return "Grout", "description"
    if "schluter color" in value:
        return "Schluter", "description"
    if "floor tile" in value:
        return "Floor Tile", "description"
    if "shower wall" in value:
        return "Shower wall", "description"
    return "", "row"


def _operation_from_entry(entry: str) -> str:
    match = re.search(r"\s-\s(ADD|DELETE|CHANGE|CLARIFICATION)\s-\s", entry, re.IGNORECASE)
    return match.group(1).upper() if match else "CLARIFICATION"


def _entry_action(entry: str, rows: list[OrderRow]) -> RevisionAction:
    operation = _operation_from_entry(entry)
    room = _room_from_entry(entry)
    item_type, target_field = _item_from_entry(entry)
    candidates = [
        (index, row)
        for index, row in enumerate(rows)
        if (not room or row.room == room)
        and (not item_type or row.item_type.casefold() == item_type.casefold())
    ]
    warnings: list[str] = []
    field_changes: dict[str, Any] = {}
    target_group = ""
    occurrence = 0
    before_value = ""
    after_value = ""
    quantity_treatment: Literal["not_applicable", "explicit", "preserved", "unchanged"] = "unchanged"

    if "inside of niche" in entry.casefold():
        after_value = "Same as Shower Wall Tile" if "wall tile" in entry.casefold() else entry
        quantity_treatment = "preserved"
        warnings.append(
            "The material allocation changed without an explicit purchasing quantity; the baseline quantity is retained."
        )
    elif operation == "CHANGE" and "corner shelf color" in entry.casefold():
        color = re.search(r"(?:to|color:)\s+([^.;]+)", entry, re.IGNORECASE)
        after_value = color.group(1).strip() if color else entry
    elif operation == "CLARIFICATION" and ":" in entry:
        after_value = entry.split(":", 1)[1].split(" - ", 1)[0].strip()
    else:
        after_value = entry

    if len(candidates) == 1:
        _index, target = candidates[0]
        target_group = _group_key(target)
        grouped = [row for row in rows if _group_key(row) == target_group]
        occurrence = grouped.index(target)
        before_value = target.description
        if operation == "CHANGE" and item_type == "Corner Shelf" and after_value:
            code_match = re.search(r"\b(CN\d+)\b\s*([^.;]*)", after_value, re.IGNORECASE)
            if code_match:
                code = code_match.group(1).upper()
                color = code_match.group(2).strip() or code
                replacement = f"{color} ({code})" if color.casefold() != code.casefold() else code
                if re.search(r"[^—-]+\(CN\d+\)\s*$", target.description, re.IGNORECASE):
                    after_value = re.sub(
                        r"[^—-]+\(CN\d+\)\s*$",
                        replacement,
                        target.description,
                        flags=re.IGNORECASE,
                    ).strip()
        if (
            target_field == "description"
            and after_value
            and operation in {"CHANGE", "CLARIFICATION"}
            and item_type in {"Grout", "Corner Shelf", "Schluter"}
        ):
            # Only apply a deterministic replacement when the correction supplies a
            # complete product value. Short semantic notes remain report-only.
            field_changes["description"] = after_value
        if (
            "inside of niche" in entry.casefold()
            and operation == "ADD"
            and after_value
        ):
            # Change orders describe a niche material swap as DELETE old
            # allocation + ADD new allocation. That changes the existing niche
            # row's material; it must not create or remove a purchasing row.
            field_changes["description"] = after_value
    elif not candidates:
        warnings.append("No generated workbook row could be matched safely.")
    else:
        warnings.append("More than one generated workbook row matches this instruction.")

    return RevisionAction(
        id=_action_id("ENTRY", entry),
        operation=operation,  # type: ignore[arg-type]
        room=room,
        item_type=item_type,
        target_field=target_field,
        before_value=before_value,
        after_value=after_value,
        evidence_corrected=entry,
        confidence="review",
        quantity_treatment=quantity_treatment,
        status="review",
        warnings=warnings,
        target_group=target_group,
        target_occurrence=occurrence,
        field_changes=field_changes,
    )


def _redact(value: str) -> str:
    value = re.sub(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", "[email]", value)
    value = re.sub(r"\b(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]\d{3}[-.\s]\d{4}\b", "[phone]", value)
    value = re.sub(
        r"\b\d{1,6}\s+[A-Za-z0-9 .'-]+(?:Street|St|Road|Rd|Drive|Dr|Lane|Ln|Court|Ct|Avenue|Ave)\b[^\n]*",
        "[address]",
        value,
        flags=re.IGNORECASE,
    )
    return value


class OpenAIInterpreter:
    def __init__(self, model: str | None = None) -> None:
        self.model = model or os.getenv("OPENAI_CORRECTION_MODEL", DEFAULT_MODEL)

    def interpret(
        self,
        entries: list[str],
        candidate_rows: list[OrderRow],
    ) -> list[LLMAction]:
        if not entries or not os.getenv("OPENAI_API_KEY"):
            return []
        try:
            from openai import OpenAI

            client = OpenAI(timeout=20.0, max_retries=1)
            candidates = [
                {
                    "index": index,
                    "room": row.room,
                    "item_type": row.item_type,
                    "description": row.description,
                }
                for index, row in enumerate(candidate_rows)
            ]
            response = client.responses.parse(
                model=self.model,
                store=False,
                input=[
                    {
                        "role": "system",
                        "content": (
                            "Interpret tile-order correction prose. Select a candidate only when it is uniquely supported. "
                            "Never calculate or estimate quantities. Copy evidence exactly from the supplied snippets."
                        ),
                    },
                    {
                        "role": "user",
                        "content": json.dumps(
                            {
                                "change_entries": [_redact(entry) for entry in entries],
                                "candidate_rows": candidates,
                            },
                            ensure_ascii=False,
                        ),
                    },
                ],
                text_format=LLMActionBatch,
            )
            parsed = response.output_parsed
            return parsed.actions if parsed else []
        except Exception:  # noqa: BLE001 - all model failures degrade to deterministic review
            return []


def _merge_llm_actions(
    deterministic: list[RevisionAction],
    llm_actions: list[LLMAction],
    rows: list[OrderRow],
) -> None:
    for item in llm_actions:
        target_group = ""
        occurrence = 0
        field_changes: dict[str, Any] = {}
        warnings: list[str] = []
        quantity_treatment: Literal["not_applicable", "explicit", "preserved", "unchanged"] = "unchanged"
        if item.affects_quantity or item.target_field in QUANTITY_FIELDS:
            quantity_treatment = "preserved"
            warnings.append(
                "The LLM identified a possible quantity impact; the baseline quantity is retained."
            )
        elif item.candidate_index is not None and 0 <= item.candidate_index < len(rows):
            target = rows[item.candidate_index]
            target_group = _group_key(target)
            grouped = [row for row in rows if _group_key(row) == target_group]
            occurrence = grouped.index(target)
            if item.target_field in ACTION_FIELDS and item.after_value:
                field_changes[item.target_field] = item.after_value
        else:
            warnings.append("The LLM did not identify one safe workbook target.")

        action = RevisionAction(
            id=_action_id("LLM", item.model_dump()),
            operation=item.operation,
            room=item.room,
            item_type=item.item_type,
            target_field=item.target_field,
            before_value=item.before_value,
            after_value=item.after_value,
            evidence_corrected=item.evidence,
            confidence="review",
            quantity_treatment=quantity_treatment,
            status="review",
            warnings=warnings,
            source="llm",
            target_group=target_group,
            target_occurrence=occurrence,
            field_changes=field_changes,
        )
        existing = next(
            (
                candidate
                for candidate in deterministic
                if candidate.evidence_corrected == action.evidence_corrected
            ),
            None,
        )
        if existing is None:
            deterministic.append(action)
            continue
        existing.source = "llm"
        existing.room = existing.room or action.room
        existing.item_type = existing.item_type or action.item_type
        existing.before_value = existing.before_value or action.before_value
        existing.after_value = action.after_value or existing.after_value
        if not existing.field_changes and action.field_changes:
            existing.field_changes = action.field_changes
            existing.target_group = action.target_group
            existing.target_occurrence = action.target_occurrence
            existing.warnings = [
                warning
                for warning in existing.warnings
                if "matched safely" not in warning and "More than one" not in warning
            ]
        if action.quantity_treatment == "preserved":
            existing.quantity_treatment = "preserved"
            existing.warnings.extend(
                warning for warning in action.warnings if warning not in existing.warnings
            )


def analyze_correction(
    original_path: Path,
    corrected_path: Path,
    template_path: Path,
    *,
    template_name: str = "classica",
    interpreter: OpenAIInterpreter | None = None,
) -> tuple[CorrectionAnalysis, ParsedCorrectionDocument, ParsedCorrectionDocument]:
    original = _parse_document(original_path, template_path)
    corrected = _parse_document(corrected_path, template_path)
    _validate_pair(original, corrected)

    actions, used_entries = _compare_rows(original.rows, corrected.rows, corrected.change_entries)
    unmatched = [entry for entry in corrected.change_entries if entry not in used_entries]
    actions.extend(_entry_action(entry, original.rows) for entry in unmatched)

    selected_interpreter = interpreter or OpenAIInterpreter()
    model = selected_interpreter.model if os.getenv("OPENAI_API_KEY") else ""
    _merge_llm_actions(
        actions,
        selected_interpreter.interpret(unmatched, original.rows),
        original.rows,
    )

    warnings: list[str] = []
    if unmatched and not os.getenv("OPENAI_API_KEY"):
        warnings.append(
            "OpenAI interpretation is unavailable; unmatched correction text requires review."
        )
    if not actions:
        warnings.append("No workbook-affecting corrections were detected.")

    analysis = CorrectionAnalysis(
        original_hash=_sha256(original_path),
        corrected_hash=_sha256(corrected_path),
        template=template_name,
        project=corrected.metadata.get("Project", ""),
        permit_number=corrected.metadata.get("Permit Number", ""),
        original_date=original.metadata.get("Order Date", ""),
        corrected_date=corrected.metadata.get("Order Date", ""),
        actions=actions,
        warnings=warnings,
        requires_review=any(action.confidence == "review" for action in actions),
        model=model,
    )
    return analysis, original, corrected


def _find_target(rows: list[OrderRow], action: RevisionAction) -> int | None:
    matching = [index for index, row in enumerate(rows) if _group_key(row) == action.target_group]
    if action.target_occurrence < len(matching):
        return matching[action.target_occurrence]
    return None


def _validated_expected_actions(
    expected: CorrectionAnalysis,
    original: ParsedCorrectionDocument,
    corrected: ParsedCorrectionDocument,
) -> list[RevisionAction]:
    deterministic, used_entries = _compare_rows(
        original.rows,
        corrected.rows,
        corrected.change_entries,
    )
    deterministic.extend(
        _entry_action(entry, original.rows)
        for entry in corrected.change_entries
        if entry not in used_entries
    )
    server_by_id = {action.id: action for action in deterministic}
    expected_by_id = {action.id: action for action in expected.actions}
    if not set(server_by_id).issubset(expected_by_id):
        raise ValueError("The correction analysis is missing document-derived actions.")

    validated: list[RevisionAction] = []
    valid_groups = {_group_key(row) for row in original.rows}
    for submitted in expected.actions:
        server_action = server_by_id.get(submitted.id)
        if server_action and submitted.source == "deterministic":
            validated.append(server_action)
            continue
        if submitted.source != "llm":
            raise ValueError("The correction analysis contains an unknown action.")
        if submitted.confidence != "review":
            raise ValueError("LLM-derived corrections must require review.")
        if submitted.evidence_corrected not in corrected.change_entries:
            raise ValueError("An LLM correction is not supported by the corrected PDF evidence.")
        if submitted.row_payload:
            raise ValueError("LLM-derived corrections cannot add workbook rows directly.")
        if not set(submitted.field_changes).issubset(ACTION_FIELDS):
            raise ValueError("An LLM correction targets an unsupported workbook field.")
        if QUANTITY_FIELDS & set(submitted.field_changes):
            raise ValueError("LLM-derived corrections cannot change purchasing quantities.")
        if submitted.field_changes and submitted.target_group not in valid_groups:
            raise ValueError("An LLM correction does not target an original workbook row.")
        submitted.status = "review"
        validated.append(submitted.model_copy(deep=True))
    return validated


def apply_actions(
    original_rows: list[OrderRow],
    actions: list[RevisionAction],
    decisions: dict[str, str],
) -> tuple[list[OrderRow], list[RevisionAction]]:
    rows = deepcopy(original_rows)
    resolved: list[RevisionAction] = []
    for source_action in actions:
        action = source_action.model_copy(deep=True)
        should_apply = action.confidence == "high"
        if action.confidence == "review":
            decision = decisions.get(action.id)
            if decision not in {"apply", "ignore"}:
                raise ValueError(f"A review decision is required for correction {action.id}.")
            should_apply = decision == "apply"
        if not should_apply:
            action.status = "ignored"
            resolved.append(action)
            continue

        target_index = _find_target(rows, action)
        if action.operation == "ADD" and action.row_payload:
            new_row = _row_from_payload(action.row_payload)
            insert_at = len(rows)
            room_indexes = [index for index, row in enumerate(rows) if row.room == new_row.room]
            if room_indexes:
                insert_at = room_indexes[-1] + 1
            rows.insert(insert_at, new_row)
            action.status = "applied"
        elif (
            action.operation == "DELETE"
            and action.target_field == "row"
            and target_index is not None
        ):
            rows.pop(target_index)
            action.status = "applied"
        elif action.field_changes and target_index is not None:
            for field, value in action.field_changes.items():
                if field in QUANTITY_FIELDS and action.quantity_treatment == "preserved":
                    continue
                setattr(rows[target_index], field, value)
            action.status = "applied"
        elif target_index is not None and action.operation in {"DELETE", "CLARIFICATION"}:
            action.status = "already_current"
        else:
            action.status = "unresolved"
            if "No deterministic workbook change could be applied." not in action.warnings:
                action.warnings.append("No deterministic workbook change could be applied.")
        resolved.append(action)
    return rows, resolved


def _append_revision_report(
    output_path: Path,
    analysis: CorrectionAnalysis,
    actions: list[RevisionAction],
) -> None:
    workbook = load_workbook(output_path)
    if "Revision Report" in workbook.sheetnames:
        del workbook["Revision Report"]
    sheet = workbook.create_sheet("Revision Report")
    sheet.merge_cells("A1:M1")
    sheet["A1"] = "REVISION REPORT"
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="204070")
    sheet["A1"].alignment = Alignment(horizontal="center")
    metadata = (
        ("Project", analysis.project),
        ("Original order date", analysis.original_date),
        ("Corrected order date", analysis.corrected_date),
        ("Model", analysis.model or "Deterministic fallback"),
        ("Prompt version", analysis.prompt_version),
    )
    row_number = 3
    for label, value in metadata:
        sheet.cell(row_number, 1, label).font = Font(bold=True)
        sheet.cell(row_number, 2, value)
        sheet.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=13)
        row_number += 1
    row_number += 1
    headers = [
        "Status", "Confidence", "Action", "Room", "Item", "Field", "Before",
        "After", "Quantity treatment", "Corrected evidence", "Warnings", "Source",
        "Action ID",
    ]
    for column, header in enumerate(headers, 1):
        sheet.cell(row_number, column, header)
    thin = Side(style="thin", color="A6A6A6")
    for cell in sheet[row_number]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E7E6E6")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_number += 1
    for action in actions:
        values = [
            action.status,
            action.confidence,
            action.operation,
            action.room,
            action.item_type,
            action.target_field,
            action.before_value,
            action.after_value,
            action.quantity_treatment,
            action.evidence_corrected,
            " ".join(action.warnings),
            action.source,
            action.id,
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_number, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row_number += 1
    widths = [14, 12, 16, 18, 20, 18, 36, 36, 20, 70, 48, 14, 18]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A10"
    sheet.auto_filter.ref = f"A9:M{max(9, row_number - 1)}"
    sheet.sheet_view.showGridLines = False
    workbook.save(output_path)


def generate_corrected_workbook(
    original_path: Path,
    corrected_path: Path,
    output_path: Path,
    template_path: Path,
    decisions: dict[str, str],
    *,
    template_name: str = "classica",
    expected_analysis: CorrectionAnalysis | None = None,
) -> tuple[CorrectionAnalysis, list[RevisionAction], list[OrderRow]]:
    if expected_analysis:
        original = _parse_document(original_path, template_path)
        corrected = _parse_document(corrected_path, template_path)
        _validate_pair(original, corrected)
        if (
            expected_analysis.original_hash != _sha256(original_path)
            or expected_analysis.corrected_hash != _sha256(corrected_path)
        ):
            raise ValueError("The uploaded PDFs do not match the analyzed documents.")
        if expected_analysis.template != template_name:
            raise ValueError("The correction analysis uses a different template.")
        if expected_analysis.project.casefold() != corrected.metadata.get("Project", "").casefold():
            raise ValueError("The correction analysis belongs to a different project.")
        analysis = expected_analysis.model_copy(deep=True)
        analysis.actions = _validated_expected_actions(analysis, original, corrected)
    else:
        analysis, original, corrected = analyze_correction(
            original_path,
            corrected_path,
            template_path,
            template_name=template_name,
        )

    rows, resolved = apply_actions(original.rows, analysis.actions, decisions)
    template = load_template(template_path)
    write_workbook(
        corrected.metadata,
        rows,
        output_path,
        corrected.display_names,
        template.get("workbook_title", "TILE ORDER"),
        template.get("header_title", "Tile Order"),
    )
    _append_revision_report(output_path, analysis, resolved)
    return analysis, resolved, rows
